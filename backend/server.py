from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import csv
import io
import uuid
import logging
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field

# ---------------- Config ----------------
JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ.get("JWT_SECRET", "dev-secret-change-me-" + uuid.uuid4().hex)
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@shift.com")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
OWNER_ROLES = {"owner", "admin"}  # keep legacy admin accounts as owners
MANAGEMENT_ROLES = {"owner", "admin", "manager"}
VALID_ROLES = {"owner", "admin", "manager", "employee"}

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ---------------- Helpers ----------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def clean_user(u: dict) -> dict:
    return {
        "id": u["id"],
        "email": u["email"],
        "name": u.get("name", ""),
        "role": u.get("role", "employee"),
        "store_location": u.get("store_location", ""),
        "created_at": u.get("created_at"),
    }


def is_owner(user: dict) -> bool:
    return user.get("role") in OWNER_ROLES


def is_manager(user: dict) -> bool:
    return user.get("role") == "manager"


def is_management(user: dict) -> bool:
    return user.get("role") in MANAGEMENT_ROLES


def scoped_store_query(user: dict) -> dict:
    if is_owner(user):
        return {}
    if is_manager(user):
        store = user.get("store_location") or ""
        if not store:
            raise HTTPException(status_code=403, detail="Manager account has no assigned store")
        return {"store_location": store}
    raise HTTPException(status_code=403, detail="Management access required")


async def get_current_user(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> dict:
    if credentials is None or not credentials.credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = credentials.credentials
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if not is_management(user):
        raise HTTPException(status_code=403, detail="Management access required")
    return user


async def require_owner(user: dict = Depends(get_current_user)) -> dict:
    if not is_owner(user):
        raise HTTPException(status_code=403, detail="Owner access required")
    return user


# ---------------- Models ----------------
class RegisterRequest(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class ShiftCreate(BaseModel):
    date: str  # YYYY-MM-DD
    start_time: str  # HH:MM
    end_time: str  # HH:MM
    note: Optional[str] = ""
    store_location: Optional[str] = ""
    shift_type: Optional[str] = ""


class CheckInRequest(BaseModel):
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    address: Optional[str] = None
    local_date: Optional[str] = None  # YYYY-MM-DD in user's local timezone
    local_time: Optional[str] = None  # HH:MM in user's local timezone


class UserRoleUpdate(BaseModel):
    role: str
    store_location: Optional[str] = ""


class AdminAttendanceCreate(BaseModel):
    user_id: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    store_location: Optional[str] = ""
    shift_id: Optional[str] = None
    check_in_local_date: Optional[str] = None
    check_in_local_time: Optional[str] = None
    check_out_local_time: Optional[str] = None
    note: Optional[str] = ""


class AdminAttendanceUpdate(BaseModel):
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    store_location: Optional[str] = None
    shift_id: Optional[str] = None
    check_in_local_date: Optional[str] = None
    check_in_local_time: Optional[str] = None
    check_out_local_time: Optional[str] = None
    note: Optional[str] = None
    approval_status: Optional[str] = None


class TaskCreate(BaseModel):
    title: str = Field(min_length=1)
    description: Optional[str] = ""
    store_location: str = Field(min_length=1)
    assigned_user_id: Optional[str] = None


class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    store_location: Optional[str] = None
    assigned_user_id: Optional[str] = None
    status: Optional[str] = None


class TaskProposalCreate(BaseModel):
    proposal_type: str
    reason: Optional[str] = ""
    proposed_title: Optional[str] = None
    proposed_description: Optional[str] = None
    proposed_assigned_user_id: Optional[str] = None


class TaskProposalReject(BaseModel):
    reason: Optional[str] = ""


# ---------------- Notifications helpers ----------------
async def _notify(user_id: str, ntype: str, title: str, body: str = "", data: Optional[dict] = None) -> None:
    if not user_id:
        return
    rec = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "type": ntype,
        "title": title,
        "body": body,
        "data": data or {},
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    try:
        await db.notifications.insert_one(rec)
    except Exception as e:
        logger.warning(f"notify failed: {e}")


def _hhmm_to_min(hhmm: str) -> Optional[int]:
    try:
        h, m = hhmm.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return None


async def swap_touches_store(req: dict, store: str) -> bool:
    if not store:
        return False
    shift_ids = [req.get("my_shift_id"), req.get("target_shift_id")]
    shift_ids = [sid for sid in shift_ids if sid]
    if not shift_ids:
        ns = req.get("new_shift") or {}
        return ns.get("store_location") == store
    count = await db.shifts.count_documents({"id": {"$in": shift_ids}, "store_location": store})
    if count > 0:
        return True
    ns = req.get("new_shift") or {}
    return ns.get("store_location") == store


async def require_store_access(user: dict, store_location: str) -> None:
    if is_owner(user):
        return
    if is_manager(user) and store_location and user.get("store_location") == store_location:
        return
    raise HTTPException(status_code=403, detail="Managers can only access their assigned store")


async def attendance_touches_store(record: dict, store_location: str) -> bool:
    if not store_location:
        return False
    if record.get("store_location") == store_location:
        return True
    if record.get("shift_id"):
        shift = await db.shifts.find_one({"id": record["shift_id"]}, {"_id": 0, "store_location": 1})
        return bool(shift and shift.get("store_location") == store_location)
    return False


async def user_has_shift_at_store(user_id: str, store_location: str) -> bool:
    if not user_id or not store_location:
        return False
    count = await db.shifts.count_documents({
        "user_id": user_id,
        "store_location": store_location,
        "approval_status": "approved",
    })
    return count > 0


def _parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid datetime")


def _duration_minutes(check_in: Optional[str], check_out: Optional[str]) -> Optional[int]:
    start = _parse_iso_datetime(check_in)
    end = _parse_iso_datetime(check_out)
    if not start or not end:
        return None
    return max(0, int((end - start).total_seconds() // 60))


# ---------------- Auth Routes ----------------
@api_router.post("/auth/register")
async def register(payload: RegisterRequest):
    email = payload.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    user = {
        "id": str(uuid.uuid4()),
        "email": email,
        "name": payload.name,
        "password_hash": hash_password(payload.password),
        "role": "employee",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user)
    token = create_access_token(user["id"], user["email"], user["role"])
    return {"token": token, "user": clean_user(user)}


@api_router.post("/auth/login")
async def login(payload: LoginRequest):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(user["id"], user["email"], user["role"])
    return {"token": token, "user": clean_user(user)}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return clean_user(user)


# ---------------- Shift Routes ----------------
@api_router.post("/shifts")
async def create_shift(payload: ShiftCreate, user: dict = Depends(get_current_user)):
    shift = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_email": user["email"],
        "user_name": user.get("name", ""),
        "date": payload.date,
        "start_time": payload.start_time,
        "end_time": payload.end_time,
        "note": payload.note or "",
        "store_location": payload.store_location or "",
        "shift_type": payload.shift_type or "",
        "status": "scheduled",
        "approval_status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.shifts.insert_one(shift)
    shift.pop("_id", None)
    # Notify owners and the manager for this store about the new shift awaiting approval.
    admin_query = {"$or": [
        {"role": {"$in": list(OWNER_ROLES)}},
        {"role": "manager", "store_location": shift["store_location"]},
    ]}
    admins = await db.users.find(admin_query, {"_id": 0, "id": 1}).to_list(50)
    for a in admins:
        await _notify(
            a["id"], "shift_pending_approval",
            f"New shift awaiting approval: {user.get('name', user['email'])}",
            f"{shift['date']} {shift['start_time']}–{shift['end_time']}",
            {"shift_id": shift["id"]},
        )
    return shift


@api_router.get("/shifts/mine")
async def my_shifts(user: dict = Depends(get_current_user)):
    shifts = await db.shifts.find({"user_id": user["id"]}, {"_id": 0}).sort("date", -1).to_list(500)
    return shifts


@api_router.patch("/shifts/{shift_id}")
async def update_my_shift(shift_id: str, payload: ShiftCreate, user: dict = Depends(get_current_user)):
    sh = await db.shifts.find_one({"id": shift_id, "user_id": user["id"]}, {"_id": 0})
    if not sh:
        raise HTTPException(status_code=404, detail="Shift not found")
    if sh.get("approval_status") == "approved":
        raise HTTPException(status_code=400, detail="Approved shifts cannot be edited. Please contact your admin.")
    update = {
        "date": payload.date,
        "start_time": payload.start_time,
        "end_time": payload.end_time,
        "note": payload.note or "",
        "store_location": payload.store_location or "",
        "shift_type": payload.shift_type or "",
        # If a previously rejected shift is edited, reset for re-review
        "approval_status": "pending",
        "rejected_reason": None,
        "rejected_at": None,
        "rejected_by": None,
    }
    await db.shifts.update_one({"id": shift_id, "user_id": user["id"]}, {"$set": update})
    new_sh = await db.shifts.find_one({"id": shift_id}, {"_id": 0})
    # Notify owners and the manager for this store so they see the updated request.
    admin_query = {"$or": [
        {"role": {"$in": list(OWNER_ROLES)}},
        {"role": "manager", "store_location": new_sh.get("store_location", "")},
    ]}
    admins = await db.users.find(admin_query, {"_id": 0, "id": 1}).to_list(50)
    for a in admins:
        await _notify(
            a["id"], "shift_pending_approval",
            f"Shift updated by {user.get('name', user['email'])} — re-approval needed",
            f"{new_sh['date']} {new_sh['start_time']}–{new_sh['end_time']}",
            {"shift_id": shift_id},
        )
    return new_sh


@api_router.delete("/shifts/{shift_id}")
async def delete_shift(shift_id: str, user: dict = Depends(get_current_user)):
    sh = await db.shifts.find_one({"id": shift_id, "user_id": user["id"]}, {"_id": 0})
    if not sh:
        raise HTTPException(status_code=404, detail="Shift not found")
    # Once a shift is approved by admin, employees cannot cancel it
    if sh.get("approval_status") == "approved":
        raise HTTPException(status_code=400, detail="Approved shifts cannot be cancelled. Please contact your admin.")
    # Prevent cancelling shifts in the past (allow today and future)
    today = datetime.now(timezone.utc).date().isoformat()
    if sh.get("date", "") < today:
        raise HTTPException(status_code=400, detail="Cannot cancel past shifts")
    await db.shifts.delete_one({"id": shift_id, "user_id": user["id"]})
    return {"ok": True}


# ---------------- Check-in Routes ----------------
@api_router.get("/attendance/status")
async def attendance_status(user: dict = Depends(get_current_user)):
    open_record = await db.attendance.find_one(
        {"user_id": user["id"], "check_out": None}, {"_id": 0}
    )
    if open_record and open_record.get("shift_id"):
        sh = await db.shifts.find_one({"id": open_record["shift_id"]}, {"_id": 0})
        if sh:
            open_record["shift_type"] = sh.get("shift_type", open_record.get("shift_type", ""))
            open_record["shift_start_time"] = sh.get("start_time", open_record.get("shift_start_time"))
            open_record["shift_end_time"] = sh.get("end_time", open_record.get("shift_end_time"))
    return {"checked_in": bool(open_record), "current": open_record}


@api_router.post("/attendance/checkin")
async def check_in(payload: CheckInRequest, user: dict = Depends(get_current_user)):
    existing = await db.attendance.find_one({"user_id": user["id"], "check_out": None})
    if existing:
        raise HTTPException(status_code=400, detail="Already checked in")

    # Late detection: try to find a scheduled shift for this user/date
    late_minutes: Optional[int] = None
    matched_shift_id: Optional[str] = None
    matched_shift_start: Optional[str] = None
    matched_shift_end: Optional[str] = None
    matched_shift_type: Optional[str] = None
    matched_shift_store: Optional[str] = None
    if payload.local_date and payload.local_time:
        sh = await db.shifts.find_one(
            {"user_id": user["id"], "date": payload.local_date},
            {"_id": 0},
        )
        if sh:
            ci_min = _hhmm_to_min(payload.local_time)
            ss_min = _hhmm_to_min(sh.get("start_time", ""))
            if ci_min is not None and ss_min is not None:
                late_minutes = max(0, ci_min - ss_min)
                matched_shift_id = sh.get("id")
                matched_shift_start = sh.get("start_time")
                matched_shift_end = sh.get("end_time")
                matched_shift_type = sh.get("shift_type")
                matched_shift_store = sh.get("store_location")

    record = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_email": user["email"],
        "user_name": user.get("name", ""),
        "check_in": datetime.now(timezone.utc).isoformat(),
        "check_out": None,
        "check_in_lat": payload.latitude,
        "check_in_lng": payload.longitude,
        "check_in_address": payload.address,
        "check_in_local_date": payload.local_date,
        "check_in_local_time": payload.local_time,
        "check_out_lat": None,
        "check_out_lng": None,
        "check_out_address": None,
        "check_out_local_time": None,
        "duration_minutes": None,
        "late_minutes": late_minutes,
        "early_leave_minutes": None,
        "approval_status": "pending",
        "store_location": matched_shift_store or "",
        "shift_id": matched_shift_id,
        "shift_start_time": matched_shift_start,
        "shift_end_time": matched_shift_end,
        "shift_type": matched_shift_type,
    }
    await db.attendance.insert_one(record)
    record.pop("_id", None)
    return record


@api_router.post("/attendance/checkout")
async def check_out(payload: CheckInRequest, user: dict = Depends(get_current_user)):
    existing = await db.attendance.find_one({"user_id": user["id"], "check_out": None})
    if not existing:
        raise HTTPException(status_code=400, detail="No active check-in")
    check_out_time = datetime.now(timezone.utc)
    check_in_time = datetime.fromisoformat(existing["check_in"])
    duration = int((check_out_time - check_in_time).total_seconds() // 60)

    # Early-leave detection vs scheduled shift end
    early_leave: Optional[int] = None
    if existing.get("shift_end_time") and payload.local_time:
        co_min = _hhmm_to_min(payload.local_time)
        se_min = _hhmm_to_min(existing["shift_end_time"])
        if co_min is not None and se_min is not None:
            early_leave = max(0, se_min - co_min)

    await db.attendance.update_one(
        {"id": existing["id"]},
        {"$set": {
            "check_out": check_out_time.isoformat(),
            "check_out_lat": payload.latitude,
            "check_out_lng": payload.longitude,
            "check_out_address": payload.address,
            "check_out_local_time": payload.local_time,
            "duration_minutes": duration,
            "early_leave_minutes": early_leave,
        }},
    )
    updated = await db.attendance.find_one({"id": existing["id"]}, {"_id": 0})
    return updated


@api_router.get("/attendance/mine")
async def my_attendance(user: dict = Depends(get_current_user)):
    records = await db.attendance.find({"user_id": user["id"]}, {"_id": 0}).sort("check_in", -1).to_list(500)
    return records


# ---------------- Admin Routes ----------------
@api_router.get("/admin/employees")
async def admin_employees(admin: dict = Depends(require_admin)):
    if is_owner(admin):
        query = {}
    else:
        store = admin.get("store_location", "")
        approved_shift_user_ids = [
            s["user_id"]
            for s in await db.shifts.find(
                {"store_location": store, "approval_status": "approved"},
                {"_id": 0, "user_id": 1},
            ).to_list(5000)
            if s.get("user_id")
        ]
        query = {"id": {"$in": sorted(set(approved_shift_user_ids))}, "role": "employee"}
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users


@api_router.patch("/admin/users/{user_id}/role")
async def admin_update_user_role(user_id: str, payload: UserRoleUpdate, _: dict = Depends(require_owner)):
    if payload.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    update = {
        "role": payload.role,
        "store_location": payload.store_location or "",
    }
    if payload.role == "manager" and not update["store_location"]:
        raise HTTPException(status_code=400, detail="Manager must be assigned to a store")
    res = await db.users.update_one({"id": user_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})


@api_router.get("/admin/attendance")
async def admin_attendance(admin: dict = Depends(require_admin)):
    if is_owner(admin):
        records = await db.attendance.find({}, {"_id": 0}).sort("check_in", -1).to_list(1000)
    else:
        store = admin.get("store_location", "")
        shift_ids = [s["id"] for s in await db.shifts.find({"store_location": store}, {"_id": 0, "id": 1}).to_list(5000)]
        records = await db.attendance.find(
            {"$or": [{"store_location": store}, {"shift_id": {"$in": shift_ids}}]},
            {"_id": 0},
        ).sort("check_in", -1).to_list(1000)
    return records


@api_router.post("/admin/attendance")
async def admin_create_attendance(payload: AdminAttendanceCreate, admin: dict = Depends(require_admin)):
    target_user = await db.users.find_one({"id": payload.user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    shift = None
    if payload.shift_id:
        shift = await db.shifts.find_one({"id": payload.shift_id}, {"_id": 0})
        if not shift:
            raise HTTPException(status_code=404, detail="Shift not found")
        if shift.get("user_id") != payload.user_id:
            raise HTTPException(status_code=400, detail="Shift does not belong to this user")

    store_location = payload.store_location or (shift or {}).get("store_location", "")
    await require_store_access(admin, store_location)

    now = datetime.now(timezone.utc).isoformat()
    check_in = payload.check_in or now
    check_out = payload.check_out
    record = {
        "id": str(uuid.uuid4()),
        "user_id": target_user["id"],
        "user_email": target_user["email"],
        "user_name": target_user.get("name", ""),
        "check_in": check_in,
        "check_out": check_out,
        "check_in_lat": None,
        "check_in_lng": None,
        "check_in_address": None,
        "check_in_local_date": payload.check_in_local_date,
        "check_in_local_time": payload.check_in_local_time,
        "check_out_lat": None,
        "check_out_lng": None,
        "check_out_address": None,
        "check_out_local_time": payload.check_out_local_time,
        "duration_minutes": _duration_minutes(check_in, check_out),
        "late_minutes": None,
        "early_leave_minutes": None,
        "approval_status": "approved",
        "approved_at": now,
        "approved_by": admin["id"],
        "store_location": store_location,
        "shift_id": payload.shift_id,
        "shift_start_time": (shift or {}).get("start_time"),
        "shift_end_time": (shift or {}).get("end_time"),
        "shift_type": (shift or {}).get("shift_type"),
        "note": payload.note or "",
        "created_by": admin["id"],
        "created_at": now,
    }
    await db.attendance.insert_one(record)
    record.pop("_id", None)
    return record


@api_router.patch("/admin/attendance/{attendance_id}")
async def admin_update_attendance(attendance_id: str, payload: AdminAttendanceUpdate, admin: dict = Depends(require_admin)):
    record = await db.attendance.find_one({"id": attendance_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Attendance not found")
    if is_manager(admin) and not await attendance_touches_store(record, admin.get("store_location", "")):
        raise HTTPException(status_code=403, detail="Managers can only edit attendance in their store")

    update = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No changes")
    if "approval_status" in update and update["approval_status"] not in {"pending", "approved", "rejected"}:
        raise HTTPException(status_code=400, detail="Invalid approval status")
    if "shift_id" in update and update["shift_id"]:
        shift = await db.shifts.find_one({"id": update["shift_id"]}, {"_id": 0})
        if not shift:
            raise HTTPException(status_code=404, detail="Shift not found")
        await require_store_access(admin, shift.get("store_location", ""))
        update["store_location"] = shift.get("store_location", "")
        update["shift_start_time"] = shift.get("start_time")
        update["shift_end_time"] = shift.get("end_time")
        update["shift_type"] = shift.get("shift_type")
    if "store_location" in update:
        await require_store_access(admin, update["store_location"])

    next_check_in = update.get("check_in", record.get("check_in"))
    next_check_out = update.get("check_out", record.get("check_out"))
    update["duration_minutes"] = _duration_minutes(next_check_in, next_check_out)
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["updated_by"] = admin["id"]
    await db.attendance.update_one({"id": attendance_id}, {"$set": update})
    return await db.attendance.find_one({"id": attendance_id}, {"_id": 0})


@api_router.post("/admin/attendance/{attendance_id}/approve")
async def admin_approve_attendance(attendance_id: str, admin: dict = Depends(require_admin)):
    record = await db.attendance.find_one({"id": attendance_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Attendance not found")
    if is_manager(admin) and not await attendance_touches_store(record, admin.get("store_location", "")):
        raise HTTPException(status_code=403, detail="Managers can only approve attendance in their store")
    await db.attendance.update_one({"id": attendance_id}, {"$set": {
        "approval_status": "approved",
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "approved_by": admin["id"],
    }})
    return await db.attendance.find_one({"id": attendance_id}, {"_id": 0})


@api_router.post("/admin/attendance/{attendance_id}/reject")
async def admin_reject_attendance(attendance_id: str, admin: dict = Depends(require_admin)):
    record = await db.attendance.find_one({"id": attendance_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Attendance not found")
    if is_manager(admin) and not await attendance_touches_store(record, admin.get("store_location", "")):
        raise HTTPException(status_code=403, detail="Managers can only reject attendance in their store")
    await db.attendance.update_one({"id": attendance_id}, {"$set": {
        "approval_status": "rejected",
        "rejected_at": datetime.now(timezone.utc).isoformat(),
        "rejected_by": admin["id"],
    }})
    return await db.attendance.find_one({"id": attendance_id}, {"_id": 0})


@api_router.get("/admin/shifts")
async def admin_shifts(admin: dict = Depends(require_admin)):
    shifts = await db.shifts.find(scoped_store_query(admin), {"_id": 0}).sort("date", -1).to_list(1000)
    return shifts


@api_router.get("/admin/stats")
async def admin_stats(admin: dict = Depends(require_admin)):
    shift_query = scoped_store_query(admin)
    total_employees = await db.users.count_documents({"role": "employee"})
    if is_owner(admin):
        active_now = await db.attendance.count_documents({"check_out": None})
    else:
        store = admin.get("store_location", "")
        shift_ids = [s["id"] for s in await db.shifts.find({"store_location": store}, {"_id": 0, "id": 1}).to_list(5000)]
        active_now = await db.attendance.count_documents({
            "check_out": None,
            "$or": [{"store_location": store}, {"shift_id": {"$in": shift_ids}}],
        })
    total_shifts = await db.shifts.count_documents(shift_query)
    today = datetime.now(timezone.utc).date().isoformat()
    shifts_today = await db.shifts.count_documents({**shift_query, "date": today})
    return {
        "total_employees": total_employees,
        "active_now": active_now,
        "total_shifts": total_shifts,
        "shifts_today": shifts_today,
    }


@api_router.get("/admin/reports")
async def admin_reports(period: str = "all", _: dict = Depends(require_owner)):
    now = datetime.now(timezone.utc)
    start: Optional[datetime] = None
    if period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)

    query: dict = {}
    if start:
        query["check_in"] = {"$gte": start.isoformat()}

    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    records = await db.attendance.find(query, {"_id": 0}).to_list(5000)

    by_user: dict = {}
    for u in users:
        by_user[u["id"]] = {
            "id": u["id"],
            "name": u.get("name", ""),
            "email": u["email"],
            "role": u.get("role", "employee"),
            "total_minutes": 0,
            "sessions": 0,
            "completed": 0,
            "active": 0,
            "last_check_in": None,
        }

    for r in records:
        uid = r.get("user_id")
        if uid not in by_user:
            continue
        b = by_user[uid]
        b["sessions"] += 1
        if r.get("check_out"):
            b["completed"] += 1
            b["total_minutes"] += int(r.get("duration_minutes") or 0)
        else:
            b["active"] += 1
        ci = r.get("check_in")
        if ci and (b["last_check_in"] is None or ci > b["last_check_in"]):
            b["last_check_in"] = ci

    rows = list(by_user.values())
    for r in rows:
        r["total_hours"] = round(r["total_minutes"] / 60, 2)
        r["avg_minutes"] = int(r["total_minutes"] / r["completed"]) if r["completed"] else 0
    rows.sort(key=lambda x: x["total_minutes"], reverse=True)

    totals = {
        "total_minutes": sum(r["total_minutes"] for r in rows),
        "total_hours": round(sum(r["total_minutes"] for r in rows) / 60, 2),
        "sessions": sum(r["sessions"] for r in rows),
        "completed": sum(r["completed"] for r in rows),
        "active": sum(r["active"] for r in rows),
        "employees_with_activity": sum(1 for r in rows if r["sessions"] > 0),
    }
    return {"period": period, "totals": totals, "rows": rows}


@api_router.get("/admin/reports/monthly")
async def admin_reports_monthly(months: int = 6, _: dict = Depends(require_owner)):
    months = max(1, min(months, 24))
    now = datetime.now(timezone.utc)
    # Compute a list of (year, month) anchors for the last `months` months (chronological)
    keys: list = []
    y, m = now.year, now.month
    for _i in range(months):
        keys.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    keys.reverse()  # oldest first

    # Earliest month start for query lower bound
    y0, m0 = keys[0]
    start_iso = datetime(y0, m0, 1, tzinfo=timezone.utc).isoformat()

    records = await db.attendance.find(
        {"check_in": {"$gte": start_iso}, "check_out": {"$ne": None}},
        {"_id": 0, "check_in": 1, "duration_minutes": 1, "user_id": 1},
    ).to_list(20000)

    buckets = {(y, m): {"minutes": 0, "sessions": 0, "users": set()} for (y, m) in keys}
    for r in records:
        try:
            d = datetime.fromisoformat(r["check_in"])
        except Exception:
            continue
        key = (d.year, d.month)
        if key in buckets:
            buckets[key]["minutes"] += int(r.get("duration_minutes") or 0)
            buckets[key]["sessions"] += 1
            buckets[key]["users"].add(r.get("user_id"))

    out = []
    for (y, m) in keys:
        b = buckets[(y, m)]
        label = datetime(y, m, 1).strftime("%b %Y")
        out.append({
            "year": y, "month": m, "label": label,
            "total_minutes": b["minutes"],
            "total_hours": round(b["minutes"] / 60, 2),
            "sessions": b["sessions"],
            "active_users": len(b["users"]),
        })
    return {"months": out}


@api_router.get("/admin/reports/{user_id}")
async def admin_employee_report(user_id: str, period: str = "all", _: dict = Depends(require_owner)):
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    now = datetime.now(timezone.utc)
    start: Optional[datetime] = None
    if period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)

    att_query: dict = {"user_id": user_id}
    shift_query: dict = {"user_id": user_id}
    if start:
        att_query["check_in"] = {"$gte": start.isoformat()}
        shift_query["created_at"] = {"$gte": start.isoformat()}

    sessions = await db.attendance.find(att_query, {"_id": 0}).sort("check_in", -1).to_list(2000)
    shifts = await db.shifts.find(shift_query, {"_id": 0}).sort("date", -1).to_list(2000)

    total_minutes = sum(int(s.get("duration_minutes") or 0) for s in sessions if s.get("check_out"))
    completed = sum(1 for s in sessions if s.get("check_out"))
    active = sum(1 for s in sessions if not s.get("check_out"))

    # Per-day breakdown (last 14 buckets in chronological order)
    by_day: dict = {}
    for s in sessions:
        if not s.get("check_in"):
            continue
        day = s["check_in"][:10]  # YYYY-MM-DD
        by_day.setdefault(day, 0)
        by_day[day] += int(s.get("duration_minutes") or 0)
    days_sorted = sorted(by_day.items(), key=lambda x: x[0])[-14:]
    daily = [{"date": d, "minutes": m} for d, m in days_sorted]

    totals = {
        "total_minutes": total_minutes,
        "total_hours": round(total_minutes / 60, 2),
        "sessions": len(sessions),
        "completed": completed,
        "active": active,
        "avg_minutes": int(total_minutes / completed) if completed else 0,
        "shifts_registered": len(shifts),
    }

    return {
        "period": period,
        "employee": {
            "id": user["id"],
            "name": user.get("name", ""),
            "email": user["email"],
            "role": user.get("role", "employee"),
        },
        "totals": totals,
        "daily": daily,
        "sessions": sessions,
        "shifts": shifts,
    }


@api_router.get("/admin/reports.csv")
async def admin_reports_export(period: str = "all", _: dict = Depends(require_owner)):
    data = await admin_reports(period=period)  # type: ignore
    rows = data["rows"]
    out = io.StringIO()
    out.write("\ufeff")  # BOM for Excel UTF-8 (Vietnamese chars)
    w = csv.writer(out)
    w.writerow(["Name", "Email", "Role", "Total Hours", "Total Minutes", "Sessions", "Completed", "Active", "Avg Minutes", "Last Check-in"])
    for r in rows:
        w.writerow([
            r.get("name", ""), r.get("email", ""), r.get("role", ""),
            r.get("total_hours", 0), r.get("total_minutes", 0),
            r.get("sessions", 0), r.get("completed", 0), r.get("active", 0),
            r.get("avg_minutes", 0), r.get("last_check_in") or "",
        ])
    filename = f"shift-report-{period}.csv"
    return Response(
        content=out.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/admin/reports/{user_id}/export.csv")
async def admin_employee_export(user_id: str, period: str = "all", _: dict = Depends(require_owner)):
    data = await admin_employee_report(user_id=user_id, period=period)  # type: ignore
    emp = data["employee"]
    sessions = data["sessions"]
    shifts = data["shifts"]

    out = io.StringIO()
    out.write("\ufeff")
    w = csv.writer(out)
    w.writerow([f"Employee: {emp.get('name','')} ({emp.get('email','')}) - period={period}"])
    w.writerow([])
    w.writerow(["Sessions"])
    w.writerow(["Date", "Check-in", "Check-out", "Duration (min)", "Duration", "Address (in)", "Lat (in)", "Lng (in)", "Address (out)"])
    for s in sessions:
        ci = s.get("check_in") or ""
        co = s.get("check_out") or ""
        dm = s.get("duration_minutes")
        date = ci[:10] if ci else ""
        dur_str = ""
        if dm is not None:
            dur_str = f"{dm // 60}h {dm % 60}m"
        w.writerow([
            date, ci, co, dm if dm is not None else "",
            dur_str,
            s.get("check_in_address", "") or "",
            s.get("check_in_lat", "") if s.get("check_in_lat") is not None else "",
            s.get("check_in_lng", "") if s.get("check_in_lng") is not None else "",
            s.get("check_out_address", "") or "",
        ])
    w.writerow([])
    w.writerow(["Registered Shifts"])
    w.writerow(["Date", "Start", "End", "Note", "Created at"])
    for s in shifts:
        w.writerow([s.get("date",""), s.get("start_time",""), s.get("end_time",""), s.get("note","") or "", s.get("created_at","")])

    safe_name = (emp.get("email") or user_id).split("@")[0]
    filename = f"shift-{safe_name}-{period}.csv"
    return Response(
        content=out.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/admin/reports.xlsx")
async def admin_reports_xlsx(period: str = "all", _: dict = Depends(require_owner)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    summary = await admin_reports(period=period)  # type: ignore

    wb = Workbook()
    # ---- "total" sheet ----
    ws = wb.active
    ws.title = "total"
    headers = ["Name", "Email", "Role", "Total Hours", "Total Minutes",
               "Sessions", "Completed", "Active", "Avg Minutes", "Last Check-in"]
    ws.append(headers)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0052FF")
    for col_idx, _h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="left")
    for r in summary["rows"]:
        ws.append([
            r.get("name", ""), r.get("email", ""), r.get("role", ""),
            r.get("total_hours", 0), r.get("total_minutes", 0),
            r.get("sessions", 0), r.get("completed", 0), r.get("active", 0),
            r.get("avg_minutes", 0), r.get("last_check_in") or "",
        ])
    # auto column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ---- one sheet per employee ----
    used_titles = {"total"}
    for r in summary["rows"]:
        detail = await admin_employee_report(user_id=r["id"], period=period)  # type: ignore
        emp = detail["employee"]
        # Excel sheet titles: max 31 chars, no \ / * ? : [ ]
        base = (emp.get("name") or emp.get("email") or "user").strip()
        for ch in "\\/*?:[]":
            base = base.replace(ch, " ")
        title = base[:28] or "user"
        suffix = 1
        unique = title
        while unique in used_titles:
            suffix += 1
            unique = f"{title[:26]} {suffix}"
        used_titles.add(unique)
        sh = wb.create_sheet(unique)

        sh["A1"] = f"{emp.get('name','')} ({emp.get('email','')})"
        sh["A1"].font = Font(bold=True, size=14)
        sh["A2"] = f"Period: {period} • Total: {detail['totals']['total_hours']}h • Sessions: {detail['totals']['sessions']} • Completed: {detail['totals']['completed']}"
        sh["A2"].font = Font(italic=True, color="666666")

        sh.append([])
        sh.append(["Sessions"])
        sh["A4"].font = Font(bold=True)
        s_headers = ["Date", "Check-in", "Check-out", "Duration (min)", "Duration",
                     "Address (in)", "Lat (in)", "Lng (in)", "Address (out)"]
        sh.append(s_headers)
        for col_idx in range(1, len(s_headers) + 1):
            c = sh.cell(row=5, column=col_idx)
            c.font = bold
            c.fill = fill
        for s in detail["sessions"]:
            ci = s.get("check_in") or ""
            co = s.get("check_out") or ""
            dm = s.get("duration_minutes")
            dur_str = f"{dm // 60}h {dm % 60}m" if dm is not None else ""
            sh.append([
                ci[:10] if ci else "",
                ci, co,
                dm if dm is not None else "",
                dur_str,
                s.get("check_in_address", "") or "",
                s.get("check_in_lat") if s.get("check_in_lat") is not None else "",
                s.get("check_in_lng") if s.get("check_in_lng") is not None else "",
                s.get("check_out_address", "") or "",
            ])

        sh.append([])
        sh.append(["Registered Shifts"])
        sh.cell(row=sh.max_row, column=1).font = Font(bold=True)
        shift_header_row = sh.max_row + 1
        sh.append(["Date", "Start", "End", "Note", "Created at"])
        for col_idx in range(1, 6):
            c = sh.cell(row=shift_header_row, column=col_idx)
            c.font = bold
            c.fill = fill
        for s in detail["shifts"]:
            sh.append([s.get("date",""), s.get("start_time",""), s.get("end_time",""),
                       s.get("note","") or "", s.get("created_at","")])

        # Auto column widths for this sheet
        for col in sh.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sh.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"shift-report-{period}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/shifts/all")
async def all_shifts(start: Optional[str] = None, end: Optional[str] = None, _: dict = Depends(get_current_user)):
    q: dict = {}
    if start and end:
        q["date"] = {"$gte": start, "$lte": end}
    shifts = await db.shifts.find(q, {"_id": 0}).sort("date", 1).to_list(2000)
    return shifts


class ShiftUpdate(BaseModel):
    date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    note: Optional[str] = None
    user_id: Optional[str] = None
    store_location: Optional[str] = None
    shift_type: Optional[str] = None


class AdminShiftCreate(ShiftCreate):
    user_id: str


@api_router.post("/admin/shifts")
async def admin_create_shift(payload: AdminShiftCreate, admin: dict = Depends(require_admin)):
    target_user = await db.users.find_one({"id": payload.user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    store_location = payload.store_location or ""
    await require_store_access(admin, store_location)
    shift = {
        "id": str(uuid.uuid4()),
        "user_id": target_user["id"],
        "user_email": target_user["email"],
        "user_name": target_user.get("name", ""),
        "date": payload.date,
        "start_time": payload.start_time,
        "end_time": payload.end_time,
        "note": payload.note or "",
        "store_location": store_location,
        "shift_type": payload.shift_type or "",
        "status": "scheduled",
        "approval_status": "approved",
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "approved_by": admin["id"],
        "created_by": admin["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.shifts.insert_one(shift)
    shift.pop("_id", None)
    await _notify(
        target_user["id"],
        "shift_assigned",
        f"New shift assigned: {shift['date']} {shift['start_time']}–{shift['end_time']}",
        f"Assigned by {admin.get('name', 'Admin')}",
        {"shift_id": shift["id"]},
    )
    return shift


@api_router.patch("/admin/shifts/{shift_id}")
async def admin_update_shift(shift_id: str, payload: ShiftUpdate, admin: dict = Depends(require_admin)):
    update = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No changes")
    old_shift = await db.shifts.find_one({"id": shift_id}, {"_id": 0})
    if not old_shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    if is_manager(admin):
        manager_store = admin.get("store_location", "")
        target_store = update.get("store_location", old_shift.get("store_location", ""))
        if old_shift.get("store_location") != manager_store or target_store != manager_store:
            raise HTTPException(status_code=403, detail="Managers can only edit shifts in their store")
    if "user_id" in update:
        new_user = await db.users.find_one({"id": update["user_id"]}, {"_id": 0})
        if not new_user:
            raise HTTPException(status_code=404, detail="User not found")
        update["user_email"] = new_user["email"]
        update["user_name"] = new_user.get("name", "")
    res = await db.shifts.update_one({"id": shift_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Shift not found")
    new_shift = await db.shifts.find_one({"id": shift_id}, {"_id": 0})

    # Notify on reassignment
    if "user_id" in update and update["user_id"] != old_shift.get("user_id"):
        await _notify(
            update["user_id"], "shift_assigned",
            f"New shift assigned: {new_shift.get('date')} {new_shift.get('start_time')}–{new_shift.get('end_time')}",
            f"Assigned by admin {admin.get('name', 'Admin')}",
            {"shift_id": shift_id},
        )
        if old_shift.get("user_id"):
            await _notify(
                old_shift["user_id"], "shift_unassigned",
                f"Shift removed: {old_shift.get('date')} {old_shift.get('start_time')}–{old_shift.get('end_time')}",
                f"Reassigned by admin {admin.get('name', 'Admin')}",
                {"shift_id": shift_id},
            )
    return new_shift


@api_router.delete("/admin/shifts/{shift_id}")
async def admin_delete_shift(shift_id: str, admin: dict = Depends(require_admin)):
    if is_manager(admin):
        sh = await db.shifts.find_one({"id": shift_id}, {"_id": 0})
        if not sh:
            raise HTTPException(status_code=404, detail="Shift not found")
        if sh.get("store_location") != admin.get("store_location", ""):
            raise HTTPException(status_code=403, detail="Managers can only delete shifts in their store")
    res = await db.shifts.delete_one({"id": shift_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Shift not found")
    return {"ok": True}


@api_router.post("/admin/shifts/{shift_id}/approve")
async def admin_approve_shift(shift_id: str, admin: dict = Depends(require_admin)):
    sh = await db.shifts.find_one({"id": shift_id}, {"_id": 0})
    if not sh:
        raise HTTPException(status_code=404, detail="Shift not found")
    if is_manager(admin) and sh.get("store_location") != admin.get("store_location", ""):
        raise HTTPException(status_code=403, detail="Managers can only approve shifts in their store")
    await db.shifts.update_one({"id": shift_id}, {"$set": {
        "approval_status": "approved",
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "approved_by": admin["id"],
        "rejected_reason": None,
    }})
    if sh.get("user_id"):
        await _notify(
            sh["user_id"], "shift_approved",
            "Your shift was approved",
            f"{sh.get('date')} {sh.get('start_time')}–{sh.get('end_time')} approved by {admin.get('name', 'Admin')}",
            {"shift_id": shift_id},
        )
    return await db.shifts.find_one({"id": shift_id}, {"_id": 0})


class ShiftRejectBody(BaseModel):
    reason: Optional[str] = ""


@api_router.post("/admin/shifts/{shift_id}/reject")
async def admin_reject_shift(shift_id: str, payload: ShiftRejectBody, admin: dict = Depends(require_admin)):
    sh = await db.shifts.find_one({"id": shift_id}, {"_id": 0})
    if not sh:
        raise HTTPException(status_code=404, detail="Shift not found")
    if is_manager(admin) and sh.get("store_location") != admin.get("store_location", ""):
        raise HTTPException(status_code=403, detail="Managers can only reject shifts in their store")
    await db.shifts.update_one({"id": shift_id}, {"$set": {
        "approval_status": "rejected",
        "rejected_at": datetime.now(timezone.utc).isoformat(),
        "rejected_by": admin["id"],
        "rejected_reason": payload.reason or "",
    }})
    if sh.get("user_id"):
        body = f"{sh.get('date')} {sh.get('start_time')}–{sh.get('end_time')} rejected by {admin.get('name', 'Admin')}"
        if payload.reason:
            body += f"\nReason: {payload.reason}"
        await _notify(
            sh["user_id"], "shift_rejected",
            "Your shift was rejected",
            body,
            {"shift_id": shift_id},
        )
    return await db.shifts.find_one({"id": shift_id}, {"_id": 0})


@api_router.get("/admin/shifts/pending")
async def admin_pending_shifts(admin: dict = Depends(require_admin)):
    query = {"approval_status": {"$in": [None, "pending"]}}
    if is_manager(admin):
        query["store_location"] = admin.get("store_location", "")
    items = await db.shifts.find(
        query,
        {"_id": 0},
    ).sort("date", 1).to_list(500)
    return items


@api_router.post("/admin/shifts/{shift_id}/unapprove")
async def admin_unapprove_shift(shift_id: str, admin: dict = Depends(require_admin)):
    sh = await db.shifts.find_one({"id": shift_id}, {"_id": 0})
    if not sh:
        raise HTTPException(status_code=404, detail="Shift not found")
    if is_manager(admin) and sh.get("store_location") != admin.get("store_location", ""):
        raise HTTPException(status_code=403, detail="Managers can only update shifts in their store")
    if sh.get("approval_status") != "approved":
        raise HTTPException(status_code=400, detail="Shift is not currently approved")
    await db.shifts.update_one({"id": shift_id}, {"$set": {
        "approval_status": "pending",
        "unapproved_at": datetime.now(timezone.utc).isoformat(),
        "unapproved_by": admin["id"],
        "approved_at": None,
        "approved_by": None,
    }})
    if sh.get("user_id"):
        await _notify(
            sh["user_id"], "shift_unapproved",
            "Approval revoked on your shift",
            f"{sh.get('date')} {sh.get('start_time')}–{sh.get('end_time')} reverted to pending by {admin.get('name', 'Admin')}",
            {"shift_id": shift_id},
        )
    return await db.shifts.find_one({"id": shift_id}, {"_id": 0})


# ---------------- Tasks ----------------
async def _validate_task_assignment(payload: TaskCreate | TaskUpdate, admin: dict, existing: Optional[dict] = None) -> tuple[str, Optional[dict]]:
    store_location = payload.store_location if payload.store_location is not None else (existing or {}).get("store_location", "")
    assigned_user_id = payload.assigned_user_id if payload.assigned_user_id is not None else (existing or {}).get("assigned_user_id")
    if not store_location:
        raise HTTPException(status_code=400, detail="Task store is required")
    await require_store_access(admin, store_location)

    assigned_user = None
    if assigned_user_id:
        assigned_user = await db.users.find_one({"id": assigned_user_id}, {"_id": 0})
        if not assigned_user:
            raise HTTPException(status_code=404, detail="Assigned user not found")
        if not await user_has_shift_at_store(assigned_user_id, store_location):
            raise HTTPException(status_code=400, detail="Assigned employee must have a shift at this store")
    elif is_manager(admin):
        raise HTTPException(status_code=400, detail="Managers must assign tasks to an employee")
    return store_location, assigned_user


@api_router.get("/admin/tasks")
async def admin_tasks(admin: dict = Depends(require_admin)):
    query = scoped_store_query(admin)
    items = await db.tasks.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.post("/admin/tasks")
async def admin_create_task(payload: TaskCreate, admin: dict = Depends(require_admin)):
    store_location, assigned_user = await _validate_task_assignment(payload, admin)
    now = datetime.now(timezone.utc).isoformat()
    task = {
        "id": str(uuid.uuid4()),
        "title": payload.title.strip(),
        "description": payload.description or "",
        "store_location": store_location,
        "assigned_user_id": payload.assigned_user_id or None,
        "assigned_user_name": (assigned_user or {}).get("name", ""),
        "assigned_user_email": (assigned_user or {}).get("email", ""),
        "status": "open",
        "created_by": admin["id"],
        "created_by_name": admin.get("name", ""),
        "created_by_role": admin.get("role", ""),
        "created_at": now,
        "completed_at": None,
        "completed_by": None,
        "completed_by_name": None,
    }
    await db.tasks.insert_one(task)
    task.pop("_id", None)

    if task["assigned_user_id"]:
        await _notify(
            task["assigned_user_id"],
            "task_assigned",
            f"New task: {task['title']}",
            f"{store_location} • assigned by {admin.get('name', 'Admin')}",
            {"task_id": task["id"]},
        )
    else:
        managers = await db.users.find(
            {"role": "manager", "store_location": store_location},
            {"_id": 0, "id": 1},
        ).to_list(50)
        for manager in managers:
            await _notify(
                manager["id"],
                "task_assigned",
                f"Store task: {task['title']}",
                f"{store_location} • assigned by {admin.get('name', 'Admin')}",
                {"task_id": task["id"]},
            )
    return task


@api_router.patch("/admin/tasks/{task_id}")
async def admin_update_task(task_id: str, payload: TaskUpdate, admin: dict = Depends(require_admin)):
    existing = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Task not found")
    await require_store_access(admin, existing.get("store_location", ""))
    if is_manager(admin) and existing.get("created_by") != admin["id"]:
        raise HTTPException(status_code=403, detail="Managers can only edit tasks they created")
    update = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No changes")
    if "status" in update and update["status"] not in {"open", "completed", "cancelled"}:
        raise HTTPException(status_code=400, detail="Invalid task status")
    if "store_location" in update or "assigned_user_id" in update:
        _, assigned_user = await _validate_task_assignment(payload, admin, existing)
        if "assigned_user_id" in update:
            update["assigned_user_name"] = (assigned_user or {}).get("name", "")
            update["assigned_user_email"] = (assigned_user or {}).get("email", "")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["updated_by"] = admin["id"]
    await db.tasks.update_one({"id": task_id}, {"$set": update})
    return await db.tasks.find_one({"id": task_id}, {"_id": 0})


@api_router.delete("/admin/tasks/{task_id}")
async def admin_delete_task(task_id: str, admin: dict = Depends(require_admin)):
    existing = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Task not found")
    await require_store_access(admin, existing.get("store_location", ""))
    if is_manager(admin) and existing.get("created_by") != admin["id"]:
        raise HTTPException(status_code=403, detail="Managers can only delete tasks they created")
    await db.tasks.delete_one({"id": task_id})
    return {"ok": True}


@api_router.get("/tasks/mine")
async def my_tasks(user: dict = Depends(get_current_user)):
    if is_owner(user):
        query = {}
    elif is_manager(user):
        query = {"store_location": user.get("store_location", ""), "$or": [
            {"assigned_user_id": None},
            {"assigned_user_id": ""},
            {"assigned_user_id": user["id"]},
        ]}
    else:
        query = {"assigned_user_id": user["id"]}
    items = await db.tasks.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api_router.post("/tasks/{task_id}/complete")
async def complete_task(task_id: str, user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    allowed = False
    if is_owner(user):
        allowed = True
    elif is_manager(user):
        allowed = (
            task.get("store_location") == user.get("store_location", "") and
            task.get("assigned_user_id") in (None, "", user["id"])
        )
    else:
        allowed = task.get("assigned_user_id") == user["id"]
    if not allowed:
        raise HTTPException(status_code=403, detail="Not allowed to complete this task")
    await db.tasks.update_one({"id": task_id}, {"$set": {
        "status": "completed",
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "completed_by": user["id"],
        "completed_by_name": user.get("name", ""),
    }})
    return await db.tasks.find_one({"id": task_id}, {"_id": 0})


def task_is_visible_to_user(task: dict, user: dict) -> bool:
    if is_owner(user):
        return True
    if is_manager(user):
        return (
            task.get("store_location") == user.get("store_location", "") and
            task.get("assigned_user_id") in (None, "", user["id"])
        )
    return task.get("assigned_user_id") == user["id"]


def can_review_task_proposal(task: dict, reviewer: dict) -> bool:
    if is_owner(reviewer):
        return True
    return (
        is_manager(reviewer) and
        task.get("store_location") == reviewer.get("store_location", "") and
        task.get("created_by") == reviewer["id"]
    )


@api_router.post("/tasks/{task_id}/proposals")
async def create_task_proposal(task_id: str, payload: TaskProposalCreate, user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if task.get("status") != "open":
        raise HTTPException(status_code=400, detail="Only open tasks can be changed")
    if not task_is_visible_to_user(task, user):
        raise HTTPException(status_code=403, detail="Not allowed to propose changes for this task")
    if payload.proposal_type not in {"cancel", "change"}:
        raise HTTPException(status_code=400, detail="Invalid proposal type")
    if payload.proposal_type == "change" and not any([
        payload.proposed_title,
        payload.proposed_description,
        payload.proposed_assigned_user_id,
    ]):
        raise HTTPException(status_code=400, detail="Provide at least one proposed change")
    existing = await db.task_proposals.find_one({"task_id": task_id, "status": "pending"}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="A pending proposal already exists for this task")

    proposed_user = None
    if payload.proposed_assigned_user_id:
        proposed_user = await db.users.find_one({"id": payload.proposed_assigned_user_id}, {"_id": 0})
        if not proposed_user:
            raise HTTPException(status_code=404, detail="Proposed assignee not found")
        if not await user_has_shift_at_store(payload.proposed_assigned_user_id, task.get("store_location", "")):
            raise HTTPException(status_code=400, detail="Proposed assignee must have a shift at this store")

    now = datetime.now(timezone.utc).isoformat()
    proposal = {
        "id": str(uuid.uuid4()),
        "task_id": task_id,
        "task_title": task.get("title", ""),
        "store_location": task.get("store_location", ""),
        "proposal_type": payload.proposal_type,
        "reason": payload.reason or "",
        "proposed_title": payload.proposed_title,
        "proposed_description": payload.proposed_description,
        "proposed_assigned_user_id": payload.proposed_assigned_user_id,
        "proposed_assigned_user_name": (proposed_user or {}).get("name", ""),
        "proposed_assigned_user_email": (proposed_user or {}).get("email", ""),
        "status": "pending",
        "requested_by": user["id"],
        "requested_by_name": user.get("name", ""),
        "requested_by_email": user.get("email", ""),
        "created_at": now,
        "reviewed_at": None,
        "reviewed_by": None,
        "reviewed_by_name": None,
        "rejected_reason": None,
    }
    await db.task_proposals.insert_one(proposal)
    proposal.pop("_id", None)

    reviewers = []
    if task.get("created_by"):
        creator = await db.users.find_one({"id": task["created_by"]}, {"_id": 0, "id": 1})
        if creator:
            reviewers.append(creator)
    owners = await db.users.find({"role": {"$in": list(OWNER_ROLES)}}, {"_id": 0, "id": 1}).to_list(50)
    reviewers.extend(owners)
    seen = set()
    for reviewer in reviewers:
        if reviewer["id"] in seen or reviewer["id"] == user["id"]:
            continue
        seen.add(reviewer["id"])
        await _notify(
            reviewer["id"],
            "task_change_requested",
            f"Task proposal: {task.get('title', '')}",
            f"{user.get('name', user['email'])} requested {payload.proposal_type}",
            {"proposal_id": proposal["id"], "task_id": task_id},
        )
    return proposal


@api_router.get("/admin/task-proposals")
async def admin_task_proposals(admin: dict = Depends(require_admin)):
    query = {} if is_owner(admin) else {"store_location": admin.get("store_location", "")}
    proposals = await db.task_proposals.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    if is_manager(admin):
        proposals = [
            p for p in proposals
            if await db.tasks.find_one({"id": p.get("task_id"), "created_by": admin["id"]}, {"_id": 0, "id": 1})
        ]
    return proposals


@api_router.post("/admin/task-proposals/{proposal_id}/approve")
async def approve_task_proposal(proposal_id: str, admin: dict = Depends(require_admin)):
    proposal = await db.task_proposals.find_one({"id": proposal_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    if proposal.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Proposal already reviewed")
    task = await db.tasks.find_one({"id": proposal.get("task_id")}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if not can_review_task_proposal(task, admin):
        raise HTTPException(status_code=403, detail="Not allowed to review this proposal")

    now = datetime.now(timezone.utc).isoformat()
    task_update = {
        "updated_at": now,
        "updated_by": admin["id"],
    }
    if proposal.get("proposal_type") == "cancel":
        task_update["status"] = "cancelled"
        task_update["cancelled_at"] = now
        task_update["cancelled_by"] = admin["id"]
    else:
        if proposal.get("proposed_title"):
            task_update["title"] = proposal["proposed_title"].strip()
        if proposal.get("proposed_description") is not None:
            task_update["description"] = proposal.get("proposed_description") or ""
        if proposal.get("proposed_assigned_user_id"):
            assigned = await db.users.find_one({"id": proposal["proposed_assigned_user_id"]}, {"_id": 0})
            if not assigned:
                raise HTTPException(status_code=404, detail="Proposed assignee not found")
            task_update["assigned_user_id"] = assigned["id"]
            task_update["assigned_user_name"] = assigned.get("name", "")
            task_update["assigned_user_email"] = assigned.get("email", "")
    await db.tasks.update_one({"id": task["id"]}, {"$set": task_update})
    await db.task_proposals.update_one({"id": proposal_id}, {"$set": {
        "status": "approved",
        "reviewed_at": now,
        "reviewed_by": admin["id"],
        "reviewed_by_name": admin.get("name", ""),
    }})
    if proposal.get("requested_by"):
        await _notify(
            proposal["requested_by"],
            "task_proposal_approved",
            "Task proposal approved",
            task.get("title", ""),
            {"proposal_id": proposal_id, "task_id": task["id"]},
        )
    return await db.task_proposals.find_one({"id": proposal_id}, {"_id": 0})


@api_router.post("/admin/task-proposals/{proposal_id}/reject")
async def reject_task_proposal(proposal_id: str, payload: TaskProposalReject, admin: dict = Depends(require_admin)):
    proposal = await db.task_proposals.find_one({"id": proposal_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    if proposal.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Proposal already reviewed")
    task = await db.tasks.find_one({"id": proposal.get("task_id")}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if not can_review_task_proposal(task, admin):
        raise HTTPException(status_code=403, detail="Not allowed to review this proposal")
    await db.task_proposals.update_one({"id": proposal_id}, {"$set": {
        "status": "rejected",
        "reviewed_at": datetime.now(timezone.utc).isoformat(),
        "reviewed_by": admin["id"],
        "reviewed_by_name": admin.get("name", ""),
        "rejected_reason": payload.reason or "",
    }})
    if proposal.get("requested_by"):
        await _notify(
            proposal["requested_by"],
            "task_proposal_rejected",
            "Task proposal rejected",
            payload.reason or task.get("title", ""),
            {"proposal_id": proposal_id, "task_id": task["id"]},
        )
    return await db.task_proposals.find_one({"id": proposal_id}, {"_id": 0})


# ---------------- Swap Requests ----------------
class NewShiftProposal(BaseModel):
    date: str
    start_time: str
    end_time: str
    note: Optional[str] = ""
    store_location: Optional[str] = ""
    shift_type: Optional[str] = ""


class SwapCreate(BaseModel):
    my_shift_id: str
    target_shift_id: Optional[str] = None
    new_shift: Optional[NewShiftProposal] = None
    message: Optional[str] = ""


@api_router.post("/swap-requests")
async def create_swap(payload: SwapCreate, user: dict = Depends(get_current_user)):
    if not payload.target_shift_id and not payload.new_shift:
        raise HTTPException(status_code=400, detail="Either target_shift_id or new_shift must be provided")
    if payload.target_shift_id and payload.new_shift:
        raise HTTPException(status_code=400, detail="Provide only one of target_shift_id or new_shift")

    mine = await db.shifts.find_one({"id": payload.my_shift_id, "user_id": user["id"]}, {"_id": 0})
    if not mine:
        raise HTTPException(status_code=404, detail="Your shift not found")

    rec_base = {
        "id": str(uuid.uuid4()),
        "from_user_id": user["id"],
        "from_user_name": user.get("name", ""),
        "from_user_email": user["email"],
        "my_shift_id": mine["id"],
        "my_shift": {"date": mine["date"], "start_time": mine["start_time"], "end_time": mine["end_time"]},
        "message": payload.message or "",
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    if payload.target_shift_id:
        # Peer swap: my_shift <-> target_shift
        target = await db.shifts.find_one({"id": payload.target_shift_id}, {"_id": 0})
        if not target:
            raise HTTPException(status_code=404, detail="Target shift not found")
        if target["user_id"] == user["id"]:
            raise HTTPException(status_code=400, detail="Cannot swap with yourself")
        existing = await db.swap_requests.find_one({
            "my_shift_id": payload.my_shift_id,
            "target_shift_id": payload.target_shift_id,
            "status": "pending",
        })
        if existing:
            raise HTTPException(status_code=400, detail="Request already pending")

        # Approved shifts require admin to accept the swap
        requires_admin = (mine.get("approval_status") == "approved" or
                          target.get("approval_status") == "approved")

        rec = {
            **rec_base,
            "kind": "peer",
            "to_user_id": target["user_id"],
            "to_user_name": target.get("user_name", ""),
            "to_user_email": target.get("user_email", ""),
            "target_shift_id": target["id"],
            "target_shift": {"date": target["date"], "start_time": target["start_time"], "end_time": target["end_time"]},
            "requires_admin": requires_admin,
        }
        await db.swap_requests.insert_one(rec)
        rec.pop("_id", None)

        # Notify target peer (always so they're aware)
        await _notify(
            target["user_id"], "swap_incoming",
            f"Swap request from {user.get('name', user['email'])}",
            f"{user.get('name', user['email'])} wants your shift on {target['date']} {target['start_time']}–{target['end_time']}",
            {"swap_id": rec["id"]},
        )
        # If requires admin, also notify admins
        if requires_admin:
            admins = await db.users.find({"$or": [
                {"role": {"$in": list(OWNER_ROLES)}},
                {"role": "manager", "store_location": target.get("store_location", "")},
                {"role": "manager", "store_location": mine.get("store_location", "")},
            ]}, {"_id": 0, "id": 1}).to_list(50)
            for a in admins:
                await _notify(
                    a["id"], "swap_pending_admin",
                    f"Swap awaiting admin approval — {user.get('name', user['email'])}",
                    "Approved shift involved — admin override required",
                    {"swap_id": rec["id"]},
                )
        return rec

    # New-shift proposal: convert my_shift's date/time/store/type into the proposed values upon admin approval
    ns = payload.new_shift
    rec = {
        **rec_base,
        "kind": "new",
        "to_user_id": None,
        "to_user_name": "",
        "to_user_email": "",
        "target_shift_id": None,
        "new_shift": ns.model_dump() if ns else None,
        "requires_admin": True,
    }
    await db.swap_requests.insert_one(rec)
    rec.pop("_id", None)
    # Notify all admins
    admins = await db.users.find({"$or": [
        {"role": {"$in": list(OWNER_ROLES)}},
        {"role": "manager", "store_location": ns.store_location if ns else ""},
        {"role": "manager", "store_location": mine.get("store_location", "")},
    ]}, {"_id": 0, "id": 1}).to_list(50)
    for a in admins:
        await _notify(
            a["id"], "swap_new_request",
            f"New-shift swap request — {user.get('name', user['email'])}",
            f"From {mine['date']} {mine['start_time']}–{mine['end_time']} to {ns.date} {ns.start_time}–{ns.end_time}",
            {"swap_id": rec["id"]},
        )
    return rec


@api_router.get("/swap-requests")
async def list_swaps(user: dict = Depends(get_current_user)):
    incoming = await db.swap_requests.find({"to_user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(200)
    outgoing = await db.swap_requests.find({"from_user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"incoming": incoming, "outgoing": outgoing}


async def _do_swap(req: dict) -> None:
    """Apply the swap (peer or new-shift) and mark accepted."""
    if req.get("kind") == "new" or not req.get("target_shift_id"):
        # Mutate my_shift in-place with new_shift values
        a = await db.shifts.find_one({"id": req["my_shift_id"]}, {"_id": 0})
        if not a:
            raise HTTPException(status_code=404, detail="Your shift no longer exists")
        ns = req.get("new_shift") or {}
        await db.shifts.update_one({"id": a["id"]}, {"$set": {
            "date": ns.get("date", a["date"]),
            "start_time": ns.get("start_time", a["start_time"]),
            "end_time": ns.get("end_time", a["end_time"]),
            "note": ns.get("note", a.get("note", "")),
            "store_location": ns.get("store_location", a.get("store_location", "")),
            "shift_type": ns.get("shift_type", a.get("shift_type", "")),
            # admin already approved this whole transaction
            "approval_status": "approved",
        }})
    else:
        a = await db.shifts.find_one({"id": req["my_shift_id"]}, {"_id": 0})
        b = await db.shifts.find_one({"id": req["target_shift_id"]}, {"_id": 0})
        if not a or not b:
            raise HTTPException(status_code=404, detail="Shifts no longer exist")
        await db.shifts.update_one({"id": a["id"]}, {"$set": {
            "user_id": b["user_id"], "user_email": b["user_email"], "user_name": b.get("user_name", ""),
        }})
        await db.shifts.update_one({"id": b["id"]}, {"$set": {
            "user_id": a["user_id"], "user_email": a["user_email"], "user_name": a.get("user_name", ""),
        }})
    await db.swap_requests.update_one({"id": req["id"]}, {"$set": {
        "status": "accepted", "resolved_at": datetime.now(timezone.utc).isoformat(),
    }})


@api_router.post("/swap-requests/{rid}/accept")
async def accept_swap(rid: str, user: dict = Depends(get_current_user)):
    req = await db.swap_requests.find_one({"id": rid}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")
    # New-shift swaps cannot be peer-accepted
    if req.get("kind") == "new" or not req.get("to_user_id"):
        raise HTTPException(status_code=403, detail="This swap requires admin approval")
    if req["to_user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Not yours to accept")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Already resolved")
    if req.get("requires_admin"):
        raise HTTPException(status_code=403, detail="This swap requires admin approval")
    await _do_swap(req)
    await _notify(
        req["from_user_id"], "swap_accepted",
        f"{user.get('name', user['email'])} accepted your swap",
        f"Your shift on {req['my_shift']['date']} has been swapped.",
        {"swap_id": rid},
    )
    return {"ok": True}


@api_router.post("/swap-requests/{rid}/reject")
async def reject_swap(rid: str, user: dict = Depends(get_current_user)):
    req = await db.swap_requests.find_one({"id": rid}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")
    if req["to_user_id"] != user["id"] and req["from_user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Not allowed")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Already resolved")
    new_status = "cancelled" if req["from_user_id"] == user["id"] else "rejected"
    await db.swap_requests.update_one({"id": rid}, {"$set": {
        "status": new_status, "resolved_at": datetime.now(timezone.utc).isoformat(),
    }})
    if new_status == "rejected":
        await _notify(
            req["from_user_id"], "swap_rejected",
            f"{user.get('name', user['email'])} rejected your swap",
            f"Shift on {req['target_shift']['date']} was not swapped.",
            {"swap_id": rid},
        )
    return {"ok": True}


# ---------------- Admin: swap override + listing ----------------
@api_router.get("/admin/swap-requests")
async def admin_list_swaps(admin: dict = Depends(require_admin)):
    items = await db.swap_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    if is_manager(admin):
        store = admin.get("store_location", "")
        filtered = []
        for item in items:
            if await swap_touches_store(item, store):
                filtered.append(item)
        items = filtered
    return items


@api_router.post("/admin/swap-requests/{rid}/force-approve")
async def admin_force_approve(rid: str, admin: dict = Depends(require_admin)):
    req = await db.swap_requests.find_one({"id": rid}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")
    if is_manager(admin) and not await swap_touches_store(req, admin.get("store_location", "")):
        raise HTTPException(status_code=403, detail="Managers can only approve swaps for their store")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Already resolved")
    await _do_swap(req)
    # Mark as admin approved
    await db.swap_requests.update_one({"id": rid}, {"$set": {
        "admin_approved": True, "admin_id": admin["id"],
    }})
    msg = f"Admin {admin.get('name', 'Admin')} approved this swap."
    await _notify(req["from_user_id"], "swap_accepted", "Swap approved by admin", msg, {"swap_id": rid})
    if req.get("to_user_id"):
        await _notify(req["to_user_id"], "swap_accepted", "Swap approved by admin", msg, {"swap_id": rid})
    return {"ok": True}


@api_router.post("/admin/swap-requests/{rid}/force-reject")
async def admin_force_reject(rid: str, admin: dict = Depends(require_admin)):
    req = await db.swap_requests.find_one({"id": rid}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")
    if is_manager(admin) and not await swap_touches_store(req, admin.get("store_location", "")):
        raise HTTPException(status_code=403, detail="Managers can only reject swaps for their store")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Already resolved")
    await db.swap_requests.update_one({"id": rid}, {"$set": {
        "status": "rejected",
        "resolved_at": datetime.now(timezone.utc).isoformat(),
        "admin_id": admin["id"],
    }})
    msg = f"Admin {admin.get('name', 'Admin')} rejected this swap."
    await _notify(req["from_user_id"], "swap_rejected", "Swap rejected by admin", msg, {"swap_id": rid})
    if req.get("to_user_id"):
        await _notify(req["to_user_id"], "swap_rejected", "Swap rejected by admin", msg, {"swap_id": rid})
    return {"ok": True}


# ---------------- Notifications API ----------------
@api_router.get("/notifications")
async def list_notifications(user: dict = Depends(get_current_user)):
    items = await db.notifications.find(
        {"user_id": user["id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    return items


@api_router.get("/notifications/unread-count")
async def unread_count(user: dict = Depends(get_current_user)):
    n = await db.notifications.count_documents({"user_id": user["id"], "read": False})
    return {"count": n}


@api_router.post("/notifications/{nid}/read")
async def mark_read(nid: str, user: dict = Depends(get_current_user)):
    res = await db.notifications.update_one(
        {"id": nid, "user_id": user["id"]}, {"$set": {"read": True}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api_router.post("/notifications/read-all")
async def mark_all_read(user: dict = Depends(get_current_user)):
    res = await db.notifications.update_many(
        {"user_id": user["id"], "read": False}, {"$set": {"read": True}}
    )
    return {"updated": res.modified_count}


@api_router.get("/")
async def root():
    return {"message": "Shift Management API"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------- Startup ----------------
@app.on_event("startup")
async def startup_event():
    await db.users.create_index("email", unique=True)
    await db.shifts.create_index("user_id")
    await db.shifts.create_index("store_location")
    await db.attendance.create_index("user_id")
    await db.attendance.create_index("store_location")
    await db.attendance.create_index("approval_status")
    await db.notifications.create_index("user_id")
    await db.tasks.create_index("store_location")
    await db.tasks.create_index("assigned_user_id")
    await db.tasks.create_index("status")
    await db.task_proposals.create_index("task_id")
    await db.task_proposals.create_index("store_location")
    await db.task_proposals.create_index("status")

    # Seed admin
    existing = await db.users.find_one({"email": ADMIN_EMAIL.lower()})
    if not existing:
        admin = {
            "id": str(uuid.uuid4()),
            "email": ADMIN_EMAIL.lower(),
            "name": "Admin",
            "password_hash": hash_password(ADMIN_PASSWORD),
            "role": "admin",
            "store_location": "",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.users.insert_one(admin)
        logger.info(f"Seeded admin: {ADMIN_EMAIL}")
    elif not verify_password(ADMIN_PASSWORD, existing["password_hash"]):
        await db.users.update_one(
            {"email": ADMIN_EMAIL.lower()},
            {"$set": {"password_hash": hash_password(ADMIN_PASSWORD), "role": "admin", "store_location": ""}},
        )
        logger.info(f"Updated admin password: {ADMIN_EMAIL}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
